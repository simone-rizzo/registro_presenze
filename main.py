import datetime
import openpyxl
from openpyxl.styles import PatternFill
import re
import random
import streamlit as st

class CalculusErrorException(Exception):
    "Raised when the the working our can't fit in the range with the max hours per day."
    pass

class StartEndDateException(Exception):
    "Raised when the start date or end date are not in the days list, maybe because they are festival days."
    pass

class SaveFileException(Exception):
    "Raised when there is an error while saving the file."
    pass


def get_cell_color(workbook, cell_coordinates):
    sheet = workbook.active
    cell_color = sheet[cell_coordinates].fill.start_color.index
    return cell_color

def set_cell_value(workbook, cell_coordinates, value):
    sheet = workbook.active
    sheet[cell_coordinates] = value

def get_cell_value(workbook, cell_coordinates):
    sheet = workbook.active
    return sheet[cell_coordinates].value

def set_cell_color(workbook, cell_coordinates, color_hex):
    sheet = workbook.active
    fill_color = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
    sheet[cell_coordinates].fill = fill_color

def set_cell_value_and_color(workbook, cell_coordinates, value, color):
    set_cell_value(workbook, cell_coordinates, value)
    set_cell_color(workbook, cell_coordinates, color)

def column_to_number(column_label):
    """Convert Excel column label to a number."""
    number = 0
    for letter in column_label:
        number = number * 26 + (ord(letter.upper()) - ord('A')) + 1
    return number

def number_to_column(n):
    """Convert a number to Excel column label."""
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def get_cells_between(start_cell, end_cell):
    """Return a list of cell coordinates between two given cells in the same row."""
    start_col_label, start_row = ''.join(filter(str.isalpha, start_cell)), ''.join(filter(str.isdigit, start_cell))
    end_col_label, end_row = ''.join(filter(str.isalpha, end_cell)), ''.join(filter(str.isdigit, end_cell))

    # Convert column labels to numbers
    start_col = column_to_number(start_col_label)
    end_col = column_to_number(end_col_label)

    # Generate list of cells between start and end cell
    cells_between = [start_cell]

    for col in range(start_col + 1, end_col):
        cell_label = number_to_column(col) + start_row
        cells_between.append(cell_label)

    cells_between.append(end_cell)
    return cells_between

def set_name(sheet, name, surname):
    sheet["B4"] = f"{name} {surname}"

def split_alpha_numeric(s):
    """Splits a string into its alphabetical and numeric parts using regex.
    return Letter, Number"""
    match = re.match(r"([a-zA-Z]+)([0-9]+)", s)
    if match:
        return match.groups()
    return None, None

def get_all_days(workbook):
    days = []
    date_cells_hash = {}
    for i, month_row in enumerate(list(row_maps.keys())):
        month_cells = get_cells_between(f"{columns[0]}{month_row}",f"{columns[1]}{month_row}")
        for d in month_cells:
            if get_cell_color(workbook, d) == WHITE_COLOR:
                mese = i+1
                lett, num = split_alpha_numeric(d)
                giorno = get_cell_value(workbook, f"{lett}{int(num)-1}")
                data = f"{giorno}/{mese}"
                # controllare se la data è inclusa nel range.
                days.append(d)
                date_cells_hash[data] = d
    return days, date_cells_hash

def set_value(workbook, coord, value, color):
    if (isinstance(value, int) and value > 0) or isinstance(value, str):
        set_cell_value_and_color(workbook, coord, value, color)
        letter, number = split_alpha_numeric(coord)
        new_coord = f"{letter}{row_maps[int(number)]}"
        set_cell_value_and_color(workbook, new_coord, value, color)

def distribute_hours(workbook, total_hours, max_per_day, days):
    # Dizionario per memorizzare le ore assegnate a ciascun giorno
    hours_per_day = {day: 0 for day in days}

    while total_hours > 0:
        # Seleziona un giorno casuale dalla lista
        day = random.choice(days)

        # Determina il massimo di ore che possono essere assegnate a quel giorno
        max_hours_for_day = min(max_per_day - hours_per_day[day], total_hours)

        # Se non si possono assegnare ore a quel giorno, continua al prossimo ciclo
        if max_hours_for_day <= 0:
            continue

        # Assegna un numero casuale di ore a quel giorno, fino al massimo
        hours_to_assign = random.randint(1, max_hours_for_day)

        # Aggiorna le ore per il giorno e le ore totali rimanenti
        hours_per_day[day] += hours_to_assign
        total_hours -= hours_to_assign

    # Imposta il valore per ogni giorno
    for day, hours in hours_per_day.items():
        set_value(workbook, day, hours, YELLOW_COLOR)

WHITE_COLOR = 0
YELLOW_COLOR = 'FFFFFF00'

row_maps = { 9:71, 13:85, 17:99, 21:113, 25:127, 29:141, 33:155, 37:169, 41:183, 45:197, 49:211, 53:225 }
columns = ["C", "AG"]

def save_uploaded_file(uploaded_file):
    try:
        with open(uploaded_file.name, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return uploaded_file.name
    except Exception as e:
        raise SaveFileException() from e

def filter_by_start_end_date(days, date_cells_hash, start_date, end_date):
    start = f"{start_date.day}/{start_date.month}"
    end = f"{end_date.day}/{end_date.month}"
    try:
        lista_chiavi = list(date_cells_hash.keys())
        start_index = lista_chiavi.index(start)
        end_index = lista_chiavi.index(end)
        return days[start_index:end_index]
    except Exception as e:
        raise StartEndDateException() from e

def filter_by_absences(workbook, days, date_cells_hash):
    for abs in st.session_state['date_list']:
        splitted = abs.split("/")
        label = splitted[-1]
        hash_date = str(int(splitted[0].split("-")[-1]))+"/"+str(int(splitted[0].split("-")[-2]))
        set_value(workbook, date_cells_hash[hash_date], label, YELLOW_COLOR)
        # remove from the list that hash
        days.remove(date_cells_hash[hash_date])
        date_cells_hash.pop(hash_date)
    return days, date_cells_hash
        

def start_function(file_path, nome, cognome, n_ore, max_per_day, start_date, end_date):
    
    # Carica il workbook
    workbook = openpyxl.load_workbook(file_path)

    # Seleziona il foglio di lavoro attivo o uno specifico
    sheet = workbook.active

    # Set name
    set_name(sheet, nome, cognome)

    # Get days
    days, date_cells_hash = get_all_days(workbook)

    days, date_cells_hash = filter_by_absences(workbook, days, date_cells_hash)

    days = filter_by_start_end_date(days, date_cells_hash, start_date, end_date)
    # filtra con data inizio e fine

    if len(days) * max_per_day < n_ore:
        # print(f"Le ore totali inserite superano le 4 ore giornaliere massime per i {len(days)} disponibili.")
        raise CalculusErrorException()

    distribute_hours(workbook, n_ore, max_per_day, days)

    filename = f"Foglio di presenza {nome}_{cognome} R&S compilato.xlsx"
    # Salva le modifiche
    workbook.save(filename)
    return filename

# Inizializza una sessione per tenere traccia delle date aggiunte
if 'date_list' not in st.session_state:
    st.session_state['date_list'] = []

# Funzione per aggiungere date all'elenco
def add_date():
    date_str = f"{selected_date}/{selected_label}"
    if date_str not in st.session_state['date_list']:
        st.session_state['date_list'].append(date_str)

# Funzione per rimuovere date dall'elenco
def remove_date(date_str):
    st.session_state['date_list'].remove(date_str)

if __name__ == "__main__":
    # workbook = openpyxl.load_workbook("FOGLIO PRESENZA 2021 R&S VUOTO.xlsx")
    # get_all_days(workbook=workbook)
    st.set_page_config(page_title="Compilatore", page_icon="📄", layout="centered")
    st.title('Compilatore foglio di Presenza')
    nome = st.text_input('Nome')
    cognome = st.text_input('Cognome')
    total_hours = st.number_input('Totale Ore', min_value=0)
    max_per_day = st.number_input('Max ore per giorno', min_value=0)
    start_date = st.date_input("Data inizio")
    end_date = st.date_input("Data fine")

    st.subheader("Inserisci Date assenze:")
    # Seleziona la data e la label
    selected_date = st.date_input("Seleziona una data")
    selected_label = st.selectbox("Seleziona una label", ["F", "P", "M", "A", "CG"])

    # Pulsante per aggiungere la data
    st.button("Aggiungi Data", on_click=add_date)

    # Mostra l'elenco delle date aggiunte
    st.subheader("Date assenze Selezionate:")
    for date in st.session_state['date_list']:
        col1, col2 = st.columns([0.8, 0.2])
        col1.write(date)
        col2.button("Rimuovi", key=date, on_click=remove_date, args=(date,))
    
    uploaded_file = st.file_uploader("Foglio di presenza Excel vuoto", type="xlsx")
    submitted = st.button("Compila foglio di presenza")

    if submitted and uploaded_file:
        try:
            file_path = save_uploaded_file(uploaded_file)
            output_file_path = start_function(file_path, nome, cognome, total_hours, max_per_day, start_date, end_date)
            if output_file_path:
                with open("./"+output_file_path, "rb") as file:
                    st.download_button(label='📥 Scarica Excel',
                                    data=file,
                                    file_name=output_file_path,
                                    mime='application/vnd.ms-excel')
        except CalculusErrorException:
            st.text(f"Il Totale ore supera il massimo di numero di ore lavorabili con {max_per_day} ore giornaliere, per il range di giorni selezionati.")
        except StartEndDateException:
            st.text("Data di inizio o fine non corretti. Potrebbero essere giorni festivi controlla meglio.")
        except SaveFileException:
            st.text("Errore nell'upload e salvataggio del file excell.")
        except Exception as e:
            print(e)
            st.error(e)
    
